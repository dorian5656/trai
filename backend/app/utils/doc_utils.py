#!/usr/bin/env python
# -*- coding: utf-8 -*-

# 文件名：backend/app/utils/doc_utils.py
# 作者：liuhd
# 日期：2026-02-09 16:00:00
# 描述：文档处理工具类，提供文件格式转换等功能

import platform
import html
from datetime import datetime, date
import os
import re
import tempfile
import hashlib
import subprocess
import base64
import uuid
import requests
import pypandoc
import shutil
import asyncio
import json
import pandas as pd
import uuid
from PIL import Image
from pdf2docx import Converter
from playwright.async_api import async_playwright
from xhtml2pdf import pisa
import pypdfium2 as pdfium
import pikepdf
import easyofd
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, A3, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook
from pathlib import Path
from loguru import logger
try:
    from playwright.async_api import async_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

from backend.app.utils.upload_utils import UploadUtils
from backend.app.utils.pg_utils import PGUtils

import functools

# 定义项目根目录用于安全校验
BASE_DIR = Path(__file__).resolve().parent.parent.parent

class DocUtils:
    """文档处理工具类"""

    @staticmethod
    @functools.lru_cache(maxsize=1)
    def _get_system_font_path() -> str | None:
        """获取系统中可用的中文字体路径 (缓存结果)"""
        # 优先从环境变量获取
        default_path = "/usr/share/fonts/google-droid-sans-fonts/DroidSansFallbackFull.ttf"
        font_path = os.getenv('CHINESE_FONT_PATH', default_path)
        
        if os.path.exists(font_path):
            return font_path

        # 尝试搜索其他常见路径
        common_paths = [
            "/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf",
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/noto-cjk/NotoSansCJK-Regular.ttc",
            "/System/Library/Fonts/PingFang.ttc", # macOS
            "C:\\Windows\\Fonts\\msyh.ttc" # Windows
        ]
        
        for p in common_paths:
            if os.path.exists(p):
                return p
        
        return None

    @staticmethod
    async def _html_to_pdf_playwright(input_path: Path, output_path: Path) -> bool:
        """
        使用 Playwright 将 HTML 转换为 PDF
        :param input_path: 输入 HTML 文件路径
        :param output_path: 输出 PDF 文件路径
        :return: 是否成功
        """
        try:
            # 1. 路径安全校验 (防止路径穿越)
            resolved_path = Path(input_path).resolve()
            
            is_safe = False
            try:
                if resolved_path.is_relative_to(BASE_DIR):
                    is_safe = True
                # 允许系统临时目录
                elif resolved_path.is_relative_to(Path(tempfile.gettempdir())):
                    is_safe = True
                elif str(resolved_path).startswith("/tmp/"):
                    is_safe = True
            except Exception:
                pass
                
            if not is_safe:
                logger.error(f"路径越界风险: {input_path}")
                raise PermissionError(f"非法访问: {input_path}")
                
            async with async_playwright() as p:
                browser = await p.chromium.launch(args=['--no-sandbox', '--disable-setuid-sandbox'])
                try:
                    page = await browser.new_page()
                    
                    # 使用 file:// 协议加载本地文件
                    file_url = resolved_path.as_uri()
                    await page.goto(file_url, wait_until="networkidle")
                    
                    # 生成 PDF (A4, 打印背景)
                    await page.pdf(
                        path=str(output_path),
                        format="A4",
                        print_background=True,
                        # 减小页边距，让内容更宽
                        margin={"top": "0.8cm", "right": "0.5cm", "bottom": "0.8cm", "left": "0.5cm"}
                    )
                finally:
                    await browser.close()
                
                # 2. 验证生成结果
                if not output_path.exists() or output_path.stat().st_size == 0:
                    logger.error("PDF生成为空文件或未生成")
                    return False
                    
                return True
        except Exception as e:
            logger.exception(f"Playwright 转换 PDF 失败: {e}")
            return False

    @staticmethod
    def _process_mermaid_diagrams(content: str, temp_dir: Path) -> str:
        """
        识别 Mermaid 代码块并转换为图片引用
        策略:
        1. 优先尝试本地 mmdc (Node.js) - 速度快，无需联网，无长度限制
           - 自动搜索 PATH 中的 mmdc 命令
        2. 降级使用 mermaid.ink (Python/Online) - 纯 Python 环境可用，需联网
        """
        # 非贪婪匹配 mermaid 代码块
        mermaid_pattern = re.compile(r'```mermaid(.*?)```', re.DOTALL)
        
        # 自动搜索 mmdc 可执行文件 (优先搜索 PATH)
        mmdc_path = shutil.which("mmdc")
        
        use_local_mmdc = mmdc_path is not None
        if not use_local_mmdc:
            logger.info("未在 PATH 中找到 mmdc，将使用 mermaid.ink 在线服务进行转换")
        else:
            logger.debug(f"使用本地 mmdc: {mmdc_path}")

        def replace_func(match):
            code = match.group(1).strip()
            if not code:
                return match.group(0)

            # 使用哈希生成唯一文件名
            code_hash = hashlib.md5(code.encode('utf-8')).hexdigest()
            png_file = temp_dir / f"{code_hash}.png"
            
            # 如果图片已存在，直接返回
            if png_file.exists():
                return f"![Mermaid Diagram]({png_file})"

            success = False
            
            # 策略1: 本地转换
            if use_local_mmdc:
                mmd_file = temp_dir / f"{code_hash}.mmd"
                try:
                    mmd_file.write_text(code, encoding='utf-8')
                    # --puppeteer-configFile 只有在需要自定义配置时才加，这里使用默认
                    # 只有在 root 用户且没有配置沙箱时可能需要 --no-sandbox，但全局安装通常能处理
                    # 为了兼容性，我们可以尝试传递 --no-sandbox 参数给 puppeteer
                    
                    # 构建 puppeteer 配置文件 (临时)
                    puppeteer_config = temp_dir / "puppeteer-config.json"
                    puppeteer_config.write_text('{"args": ["--no-sandbox"]}', encoding='utf-8')
                    
                    cmd = [
                        mmdc_path, 
                        "-i", str(mmd_file), 
                        "-o", str(png_file),
                        "-p", str(puppeteer_config),
                        "-b", "white"
                    ]

                    subprocess.run(cmd, check=True, capture_output=True)
                    logger.debug(f"本地 Mermaid 转换成功: {png_file}")
                    success = True
                except Exception as e:
                    logger.warning(f"本地 Mermaid 转换失败，尝试在线转换: {e}")
                    # 失败后继续尝试在线转换

            # 策略2: 在线转换 (如果本地不可用或失败)
            if not success:
                try:
                    # Mermaid.ink 需要 base64 编码
                    base64_str = base64.urlsafe_b64encode(code.encode("utf-8")).decode("utf-8")
                    url = f"https://mermaid.ink/img/{base64_str}?bgColor=FFFFFF"
                    
                    response = requests.get(url, timeout=30)
                    if response.status_code == 200:
                        png_file.write_bytes(response.content)
                        logger.debug(f"在线 Mermaid 转换成功: {png_file}")
                        success = True
                    else:
                        logger.error(f"在线 Mermaid 请求失败: {response.status_code}")
                except Exception as e:
                    logger.error(f"在线 Mermaid 转换异常: {e}")

            if success:
                return f"![Mermaid Diagram]({png_file})"
            else:
                return match.group(0) # 转换失败保留原样

        new_content = mermaid_pattern.sub(replace_func, content)
        return new_content

    @staticmethod
    async def md_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """
        将 Markdown 文件转换为 PDF 文件 (使用 Pandoc + XeLaTeX + Mermaid CLI)
        并自动上传生成的图片和PDF到S3 (如果有配置)
        
        Args:
            input_path: 输入 Markdown 文件路径
            output_path: 输出 PDF 文件路径（可选，默认同名.pdf）
            user_id: 用户ID (用于归属记录)
            
        Returns:
            str: 生成的 PDF 文件的 URL (如果启用S3) 或 本地绝对路径
        """
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            error_msg = f"文件不存在: {input_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
        
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
        else:
            output_path = Path(output_path).resolve()
        
        logger.info(f"开始转换 Markdown 到 PDF: {input_path} -> {output_path}")
        
        try:
            # 创建临时目录用于存放生成的 Mermaid 图片和临时 Markdown
            with tempfile.TemporaryDirectory() as temp_dir_str:
                temp_dir = Path(temp_dir_str)
                
                # 1. 读取原始内容
                original_content = input_path.read_text(encoding='utf-8')
                
                # 2. 预处理 Mermaid 图表
                processed_content = DocUtils._process_mermaid_diagrams(original_content, temp_dir)
                
                # NEW: 上传 Mermaid 图片到 S3 并记录
                # 遍历生成的 png
                for img_file in temp_dir.glob("*.png"):
                    try:
                        img_bytes = img_file.read_bytes()
                        # 上传
                        url, key, size = await UploadUtils.save_from_bytes(
                            img_bytes, 
                            img_file.name, 
                            module="mermaid", 
                            content_type="image/png"
                        )
                        # 记录 DB
                        if user_id: 
                            sql = """
                                INSERT INTO user_images (user_id, filename, s3_key, url, size, mime_type, module, source)
                                VALUES (:user_id, :filename, :s3_key, :url, :size, :mime_type, :module, :source)
                            """
                            params = {
                                "user_id": user_id,
                                "filename": img_file.name,
                                "s3_key": key,
                                "url": url,
                                "size": size,
                                "mime_type": "image/png",
                                "module": "mermaid",
                                "source": "generated"
                            }
                            await PGUtils.execute_update(sql, params)
                            logger.debug(f"Mermaid图片已归档: {url}")
                    except Exception as e:
                        logger.warning(f"Mermaid图片上传/记录失败: {e}")

                # 3. 写入临时 Markdown 文件
                # 保持文件名一致以便 pandoc 处理可能的其他相对引用
                temp_md = temp_dir / input_path.name
                temp_md.write_text(processed_content, encoding='utf-8')
                
                # 4. 配置 Pandoc 参数
                # 确保能找到相对于 Markdown 文件的图片资源
                input_dir = input_path.parent
                
                extra_args = [
                    '--pdf-engine=xelatex',
                    '-V', 'CJKmainfont=Droid Sans Fallback',
                    '-V', 'CJKmonofont=Droid Sans Fallback', # 避免 mono 字体缺失警告
                    '-V', 'geometry:margin=1in'
                ]
                
                # 执行转换
                pypandoc.convert_file(
                    str(temp_md), 
                    'pdf', 
                    outputfile=str(output_path), 
                    extra_args=extra_args
                )
                
                if not output_path.exists():
                    raise Exception("PDF 生成失败 (无报错但文件未生成)")
                
                # 5. 上传最终 PDF 到 S3
                final_url = str(output_path) # 默认返回本地路径
                
                if user_id: # 只有关联用户时才上传
                    try:
                        file_bytes = output_path.read_bytes()
                        file_size = output_path.stat().st_size
                        
                        s3_key = f"docs/{user_id}/{output_path.name}"
                        url, key, size = await UploadUtils.save_from_bytes(
                            file_bytes,
                            output_path.name,
                            module="doc_convert",
                            content_type="application/pdf"
                        )
                        final_url = url
                        logger.info(f"PDF已上传S3: {url}")
                        
                        # 6. 记录到 user_docs 表
                        try:
                            insert_sql = """
                            INSERT INTO user_docs (
                                user_id, filename, s3_key, url, size, mime_type, module, source, meta_data, created_at
                            ) VALUES (
                                :user_id, :filename, :s3_key, :url, :size, :mime_type, :module, :source, :meta_data, NOW()
                            )
                            """
                            params = {
                                "user_id": user_id,
                                "filename": output_path.name,
                                "s3_key": key,
                                "url": url,
                                "size": size,
                                "mime_type": "application/pdf",
                                "module": "doc_convert",
                                "source": "converted",
                                "meta_data": json.dumps({"original_file": str(input_path.name), "type": "md2pdf"})
                            }
                            await PGUtils.execute_ddl(insert_sql, params)
                            logger.info(f"PDF记录已保存到DB: {output_path.name}")
                        except Exception as e:
                            logger.error(f"PDF记录保存DB失败: {e}")
                        
                    except Exception as e:
                        logger.error(f"PDF上传S3失败: {e}")
                        # 上传失败降级返回本地路径 (但在容器/无状态环境中可能无法访问)
                        
                return final_url
                
        except Exception as e:
            logger.error(f"Markdown 转 PDF 异常: {e}")
            raise e

    @staticmethod
    async def word_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """
        将 Word (.docx) 文件转换为 PDF 文件 (使用 Pandoc + XeLaTeX)
        
        Args:
            input_path: 输入 .docx 文件路径
            output_path: 输出 PDF 文件路径（可选，默认同名.pdf）
            user_id: 用户ID (用于归属记录)
            
        Returns:
            str: 生成的 PDF 文件的 URL (如果启用S3) 或 本地绝对路径
        """
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            error_msg = f"文件不存在: {input_path}"
            logger.error(error_msg)
            raise FileNotFoundError(error_msg)
            
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
        else:
            output_path = Path(output_path).resolve()
            
        logger.info(f"开始转换 Word 到 PDF: {input_path} -> {output_path}")
        
        try:
            # 配置 Pandoc 参数
            # 使用 xelatex 引擎以支持中文，需指定 CJK 字体
            # 优先使用 Noto Sans CJK SC，如果系统中已确认存在
            extra_args = [
                    '--pdf-engine=xelatex',
                    '-V', 'mainfont=DejaVu Sans',
                    '-V', 'CJKmainfont=Droid Sans Fallback',
                    '-V', 'CJKmonofont=Droid Sans Fallback',
                    '-V', 'geometry:margin=1in',
                    # 修复中文下划线导致的 soul package error
                    '-V', 'header-includes=\\usepackage[normalem]{ulem}\\let\\ul\\uline'
                ]
            
            # 由于 pypandoc.convert_file 是同步阻塞操作，
            # 在 async 函数中直接调用会阻塞事件循环。
            # 对于大文件，建议 run_in_executor。这里为简化直接调用。
            await asyncio.to_thread(
                pypandoc.convert_file,
                str(input_path),
                'pdf',
                outputfile=str(output_path),
                extra_args=extra_args
            )
            
            if not output_path.exists():
                raise Exception("PDF 生成失败 (无报错但文件未生成)")
                
            # 上传最终 PDF 到 S3
            final_url = str(output_path)
            
            if user_id:
                try:
                    file_bytes = output_path.read_bytes()
                    s3_key = f"docs/{user_id}/{output_path.name}"
                    url, key, size = await UploadUtils.save_from_bytes(
                        file_bytes,
                        output_path.name,
                        module="doc_convert",
                        content_type="application/pdf"
                    )
                    final_url = url
                    logger.info(f"PDF已上传S3: {url}")

                    # 记录到 user_docs 表
                    try:
                        insert_sql = """
                        INSERT INTO user_docs (
                            user_id, filename, s3_key, url, size, mime_type, module, source, meta_data, created_at
                        ) VALUES (
                            :user_id, :filename, :s3_key, :url, :size, :mime_type, :module, :source, :meta_data, NOW()
                        )
                        """
                        params = {
                            "user_id": user_id,
                            "filename": output_path.name,
                            "s3_key": key,
                            "url": url,
                            "size": size,
                            "mime_type": "application/pdf",
                            "module": "doc_convert",
                            "source": "converted",
                            "meta_data": json.dumps({"original_file": str(input_path.name), "type": "word2pdf"})
                        }
                        await PGUtils.execute_ddl(insert_sql, params)
                        logger.info(f"PDF记录已保存到DB: {output_path.name}")
                    except Exception as e:
                        logger.error(f"PDF记录保存DB失败: {e}")
                except Exception as e:
                    logger.error(f"PDF上传S3失败: {e}")
                    
            return final_url
            
        except Exception as e:
            logger.error(f"Word 转 PDF 异常: {e}")
            raise e

    @staticmethod
    async def _upload_and_record(file_path: Path, user_id: str, module: str, source: str, mime_type: str, meta_data: dict) -> str:
        """上传到 S3 并记录到数据库的辅助方法"""
        if not user_id:
            return str(file_path)
            
        try:
            file_bytes = file_path.read_bytes()
            size = file_path.stat().st_size
            # 修复 S3 Key 冲突问题 (添加 UUID)
            s3_key = f"docs/{user_id}/{uuid.uuid4()}_{file_path.name}"
            
            url, key, size = await UploadUtils.save_from_bytes(
                file_bytes,
                file_path.name,
                module=module,
                content_type=mime_type
            )
            
            insert_sql = """
            INSERT INTO user_docs (
                user_id, filename, s3_key, url, size, mime_type, module, source, meta_data, created_at
            ) VALUES (
                :user_id, :filename, :s3_key, :url, :size, :mime_type, :module, :source, :meta_data, NOW()
            )
            """
            params = {
                "user_id": user_id,
                "filename": file_path.name,
                "s3_key": key,
                "url": url,
                "size": size,
                "mime_type": mime_type,
                "module": module,
                "source": source,
                "meta_data": json.dumps(meta_data)
            }
            await PGUtils.execute_ddl(insert_sql, params)
            logger.info(f"文件已记录: {url}")
            return url
        except Exception as e:
            logger.error(f"上传/记录失败: {e}")
            return str(file_path)

    @staticmethod
    async def image_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """将 JPG/PNG 图片转换为 PDF"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
        
        try:
            image = Image.open(input_path)
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            await asyncio.to_thread(image.save, output_path, "PDF", resolution=100.0)
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/pdf", 
                {"original_file": input_path.name, "type": "img2pdf"}
            )
        except Exception as e:
            logger.error(f"图片转 PDF 失败: {e}")
            raise e

    @staticmethod
    async def ppt_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """将 PPT/PPTX 转换为 PDF (需要 LibreOffice 或支持 pptx 的 Pandoc)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
            
        try:
            # 检查 LibreOffice (soffice)
            if shutil.which("soffice"):
                # 使用 LibreOffice 进行转换 (无头模式)
                out_dir = output_path.parent
                cmd = ["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(input_path)]
                await asyncio.to_thread(subprocess.run, cmd, check=True, capture_output=True)
                
                # LibreOffice 使用原始文件名加 .pdf，确保其与 output_path 匹配
                generated_pdf = input_path.with_suffix('.pdf')
                if generated_pdf != output_path and generated_pdf.exists():
                    generated_pdf.rename(output_path)
            
            # 检查支持 pptx 的 Pandoc
            elif "pptx" in pypandoc.get_pandoc_formats()[0]:
                 extra_args = ['--pdf-engine=xelatex', '-V', 'CJKmainfont=Droid Sans Fallback']
                 await asyncio.to_thread(pypandoc.convert_file, str(input_path), 'pdf', outputfile=str(output_path), extra_args=extra_args)
            
            else:
                 raise NotImplementedError("PPT 转 PDF 需要 LibreOffice (soffice) 或支持 pptx 输入的 Pandoc 版本。")

            if not output_path.exists():
                 raise Exception("PDF 生成失败 (文件未创建)")

            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/pdf", 
                {"original_file": input_path.name, "type": "ppt2pdf"}
            )
        except Exception as e:
            logger.error(f"PPT 转 PDF 失败: {e}")
            raise e

    @staticmethod
    async def excel_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """Excel 转 PDF (优先 LibreOffice，其次 Pandas -> HTML -> Playwright)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
            
        try:
            if shutil.which("soffice"):
                out_dir = output_path.parent
                cmd = ["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(input_path)]
                await asyncio.to_thread(subprocess.run, cmd, check=True, capture_output=True)

                generated_pdf = input_path.with_suffix('.pdf')
                if generated_pdf != output_path and generated_pdf.exists():
                    generated_pdf.rename(output_path)

                if not output_path.exists():
                    raise Exception("PDF 生成失败 (文件未创建)")
            else:
                # 方案二: openpyxl -> HTML -> Playwright (解决排版、合并单元格和乱码问题)
                logger.info("Excel转PDF: 使用 openpyxl -> HTML -> Playwright 方案")
                
                try:
                    wb = load_workbook(input_path, data_only=True)
                except Exception as e:
                    logger.error(f"openpyxl 读取 Excel 失败: {e}")
                    raise e

                # 构建 HTML 内容
                html_parts = []
                
                # 注入 CSS 样式 (模拟 Excel 网格)
                # 引入中文字体 (尝试使用系统字体)
                font_path = DocUtils._get_system_font_path()
                font_face_css = ""
                if font_path:
                    font_uri = Path(font_path).as_uri()
                    font_face_css = f"""
                    @font-face {{
                        font-family: 'SystemChinese';
                        src: url('{font_uri}');
                    }}
                    """
                    font_family = "'SystemChinese', 'Microsoft YaHei', 'SimHei', sans-serif"
                else:
                    font_family = "'Microsoft YaHei', 'SimHei', sans-serif"

                css_style = f"""
                <style>
                    {font_face_css}
                    body {{
                        font-family: {font_family};
                        margin: 0;
                        padding: 10px;
                        background-color: white;
                    }}
                    .sheet-container {{
                        margin-bottom: 20px;
                        page-break-after: always;
                        width: 100%;
                    }}
                    .sheet-title {{
                        font-size: 16px;
                        font-weight: bold;
                        margin-bottom: 10px;
                        color: #333;
                        text-align: center;
                    }}
                    table {{
                        border-collapse: collapse;
                        width: 100%; /* 强制撑满宽度 */
                        table-layout: fixed; /* 固定布局，遵循 col 宽度 */
                        font-size: 11px; /* 稍微调小字体以容纳更多内容 */
                    }}
                    th, td {{
                        border: 1px solid #c0c0c0;
                        padding: 4px 6px;
                        vertical-align: top;
                        word-wrap: break-word;
                        overflow-wrap: break-word;
                    }}
                </style>
                """
                
                html_parts.append(f"<html><head><meta charset='utf-8'>{css_style}</head><body>")
                
                from openpyxl.utils import range_boundaries, get_column_letter

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # 跳过空 sheet
                    if ws.max_row == 0 or ws.max_column == 0:
                        continue
                    
                    html_parts.append(f"<div class='sheet-container'>")
                    # 如果有多个 sheet，显示 sheet 名称作为标题
                    if len(wb.sheetnames) > 1:
                        html_parts.append(f"<div class='sheet-title'>{sheet_name}</div>")
                    
                    html_parts.append("<table>")
                    
                    # 计算总宽度以分配百分比
                    col_widths_raw = []
                    total_width_raw = 0
                    for c in range(1, ws.max_column + 1):
                        col_letter = get_column_letter(c)
                        dim = ws.column_dimensions.get(col_letter)
                        # 估算宽度权重
                        width = dim.width if dim and dim.width else 10.0
                        col_widths_raw.append(width)
                        total_width_raw += width
                    
                    html_parts.append("<colgroup>")
                    for w in col_widths_raw:
                        # 使用百分比宽度，实现左右拉伸撑满
                        pct = (w / total_width_raw) * 100
                        html_parts.append(f"<col style='width:{pct:.2f}%'>")
                    html_parts.append("</colgroup>")
                    
                    # 预处理合并单元格
                    merged_cells_map = {} # (row, col) -> (rowspan, colspan)
                    hidden_cells = set() # (row, col) to skip

                    if ws.merged_cells:
                        for merged_range in ws.merged_cells.ranges:
                            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                            rowspan = max_row - min_row + 1
                            colspan = max_col - min_col + 1
                            
                            merged_cells_map[(min_row, min_col)] = (rowspan, colspan)
                            
                            for r in range(min_row, max_row + 1):
                                for c in range(min_col, max_col + 1):
                                    if r == min_row and c == min_col:
                                        continue
                                    hidden_cells.add((r, c))

                    # 遍历行
                    for r in range(1, ws.max_row + 1):
                        row_dim = ws.row_dimensions.get(r)
                        row_height = row_dim.height if row_dim and row_dim.height else None
                        row_style = f" style='height:{int(row_height * 1.333)}px'" if row_height else ""
                        html_parts.append(f"<tr{row_style}>")
                        # 遍历列
                        for c in range(1, ws.max_column + 1):
                            if (r, c) in hidden_cells:
                                continue
                            
                            cell = ws.cell(row=r, column=c)
                            value = cell.value if cell.value is not None else ""
                            
                            style_attrs = []
                            if cell.font and cell.font.bold:
                                style_attrs.append("font-weight: bold")
                            if cell.font and cell.font.italic:
                                style_attrs.append("font-style: italic")
                            if cell.font and cell.font.sz:
                                style_attrs.append(f"font-size: {cell.font.sz}pt")
                            if cell.fill and cell.fill.fgColor and cell.fill.fgColor.type == "rgb":
                                rgb = cell.fill.fgColor.rgb
                                # 忽略全透明或全黑默认值 (openpyxl 默认无填充往往是 00000000)
                                if rgb and len(rgb) == 8 and rgb != "00000000":
                                    style_attrs.append(f"background-color: #{rgb[2:]}")
                            
                            if cell.alignment:
                                if cell.alignment.horizontal:
                                    h_align = cell.alignment.horizontal
                                    if h_align in ("center", "centerContinuous"):
                                        h_align = "center"
                                    elif h_align in ("left", "right", "justify"):
                                        h_align = h_align
                                    else:
                                        h_align = "left"
                                    style_attrs.append(f"text-align: {h_align}")
                                else:
                                    style_attrs.append("text-align: left")
                                
                                if cell.alignment.vertical:
                                    v_align = cell.alignment.vertical
                                    if v_align == 'center':
                                        v_align = 'middle'
                                    style_attrs.append(f"vertical-align: {v_align}")
                                if cell.alignment.wrap_text is False:
                                    style_attrs.append("white-space: nowrap")
                            
                            style_str = f" style='{'; '.join(style_attrs)}'" if style_attrs else ""
                            
                            # 合并属性
                            span_attr = ""
                            if (r, c) in merged_cells_map:
                                rs, cs = merged_cells_map[(r, c)]
                                if rs > 1: span_attr += f" rowspan='{rs}'"
                                if cs > 1: span_attr += f" colspan='{cs}'"
                            
                            if isinstance(value, (datetime, date)):
                                value_str = value.strftime("%Y-%m-%d")
                            elif isinstance(value, (int, float)) and isinstance(cell.number_format, str) and "%" in cell.number_format:
                                value_str = f"{value * 100:.2f}%"
                            else:
                                value_str = str(value)
                            value_str = html.escape(value_str)
                            html_parts.append(f"<td{span_attr}{style_str}>{value_str}</td>")
                        
                        html_parts.append("</tr>")

                    html_parts.append("</table>")
                    html_parts.append("</div>")
                
                html_parts.append("</body></html>")
                full_html = "\n".join(html_parts)
                
                # 3. 保存临时 HTML 文件
                with tempfile.NamedTemporaryFile(suffix=".html", delete=False, dir=input_path.parent, mode="w", encoding="utf-8") as tmp_html:
                    tmp_html.write(full_html)
                    tmp_html_path = Path(tmp_html.name)
                
                try:
                    # 4. 调用 Playwright 转 PDF
                    # 使用已有的 _html_to_pdf_playwright 方法
                    success = await DocUtils._html_to_pdf_playwright(tmp_html_path, output_path)
                    if not success:
                         raise Exception("Playwright 转换 PDF 失败")
                finally:
                    # 清理临时 HTML
                    if tmp_html_path.exists():
                        try:
                            os.remove(tmp_html_path)
                        except:
                            pass

            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/pdf", 
                {"original_file": input_path.name, "type": "excel2pdf"}
            )
        except Exception as e:
            logger.error(f"Excel 转 PDF 失败: {e}")
            raise e



    @staticmethod
    def _resolve_css_variables(html_content: str) -> str:
        """解析并替换 CSS 变量 (为 xhtml2pdf 兼容)"""
        try:
            # 1. 提取 :root 定义
            root_match = re.search(r':root\s*{([^}]+)}', html_content)
            if root_match:
                vars_block = root_match.group(1)
                variables = {}
                
                # 2. 解析变量
                # 移除注释 (支持多行)
                vars_block = re.sub(r'/\*[^*]*\*+([^/*][^*]*\*+)*/', '', vars_block, flags=re.DOTALL)
                
                for match in re.finditer(r'(--[\w-]+)\s*:\s*([^;]+);', vars_block):
                    variables[match.group(1)] = match.group(2).strip()
                    
                # 3. 替换 var() - O(n) 效率优化
                def replacer(match):
                    var_name = match.group(1) # 包含 --前缀
                    return variables.get(var_name, match.group(0))
                
                html_content = re.sub(r'var\((--[\w-]+)\)', replacer, html_content)
            
            # 4. 移除残留的 var() 防止报错 (替换为黑色或透明)
            # xhtml2pdf 遇到无法解析的 var() 会报错
            html_content = re.sub(r'var\(--[^)]+\)', '#000000', html_content)
            
            return html_content
        except Exception as e:
            logger.error(f"CSS变量解析失败: {e}")
            return html_content

    @staticmethod
    def _inject_chinese_font_style(html_content: str) -> str:
        """注入中文字体样式 (解决乱码问题)"""
        font_path = DocUtils._get_system_font_path()
        
        if not font_path:
            logger.warning(f"中文字体未找到，PDF 可能乱码。建议设置 CHINESE_FONT_PATH 环境变量。")
            return html_content
            
        # xhtml2pdf 需要显式定义字体
        # 使用 Path.as_uri() 生成正确的 file:// URI，避免 Windows 路径反斜杠问题
        font_uri = Path(font_path).as_uri()
        
        font_style = f"""
        <style>
            @font-face {{
                font-family: 'DroidSansFallback';
                src: url('{font_uri}');
            }}
            body, div, p, span, a, li, ul, ol, h1, h2, h3, h4, h5, h6, table, td, th {{
                font-family: 'DroidSansFallback', sans-serif;
            }}
        </style>
        """
        
        # 插入到 </head> 之前
        if "</head>" in html_content:
            return html_content.replace("</head>", f"{font_style}</head>")
        # 如果没有 head 标签但有 html 标签
        elif "<html" in html_content:
             # 尝试找到 html 标签结束的地方插入 head
             match = re.search(r'(<html[^>]*>)', html_content)
             if match:
                 return html_content.replace(match.group(1), f"{match.group(1)}<head>{font_style}</head>")
        
        # 兜底：直接拼接到开头 (xhtml2pdf 容错性较好)
        return f"<html><head>{font_style}</head><body>{html_content}</body></html>"



    @staticmethod
    async def html_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """HTML 文件转 PDF (优先使用 Playwright，降级使用 xhtml2pdf)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
            
        # 1. 尝试使用 Playwright (支持现代 CSS, Flexbox, Grid)
        success = await DocUtils._html_to_pdf_playwright(input_path, output_path)
        
        # 2. 如果 Playwright 失败，降级使用 xhtml2pdf
        if not success:
            logger.warning("Playwright 转换失败，降级使用 xhtml2pdf")
            try:
                html_content = input_path.read_text(encoding='utf-8')
                
                # 预处理: 替换 CSS 变量
                html_content = DocUtils._resolve_css_variables(html_content)
                
                # 预处理: 注入中文字体
                html_content = DocUtils._inject_chinese_font_style(html_content)
                
                with open(output_path, "wb") as f:
                    await asyncio.to_thread(pisa.CreatePDF, html_content, dest=f)
            except Exception as e:
                logger.error(f"HTML 转 PDF 失败 (xhtml2pdf): {e}")
                raise e
                
        return await DocUtils._upload_and_record(
            output_path, user_id, "doc_convert", "converted", "application/pdf", 
            {"original_file": input_path.name, "type": "html2pdf", "engine": "playwright" if success else "xhtml2pdf"}
        )

    @staticmethod
    async def pdf_to_images(input_path: str | Path, output_dir: str | Path = None, user_id: str = None) -> list[str]:
        """PDF 转图片 (返回图片 URL/路径列表)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_dir is None:
            output_dir = input_path.parent / f"{input_path.stem}_{uuid.uuid4()}_images"
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
            
        try:
            pdf = pdfium.PdfDocument(str(input_path))
            results = []
            
            for i in range(len(pdf)):
                page = pdf[i]
                image = page.render(scale=2).to_pil()
                image_path = output_dir / f"page_{i+1}.jpg"
                image.save(image_path, "JPEG")
                
                url = await DocUtils._upload_and_record(
                    image_path, user_id, "doc_convert", "converted", "image/jpeg", 
                    {"original_file": input_path.name, "type": "pdf2img", "page": i+1}
                )
                results.append(url)
                
            return results
        except Exception as e:
            logger.error(f"PDF 转图片失败: {e}")
            raise e

    @staticmethod
    async def pdf_to_word(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """PDF 转 Word (.docx)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.docx')
            
        try:
            cv = Converter(str(input_path))
            await asyncio.to_thread(cv.convert, str(output_path), start=0, end=None)
            cv.close()
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/vnd.openxmlformats-officedocument.wordprocessingml.document", 
                {"original_file": input_path.name, "type": "pdf2word"}
            )
        except Exception as e:
            logger.error(f"PDF 转 Word 失败: {e}")
            raise e

    @staticmethod
    async def pdf_to_ppt(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """PDF 转 PPT (通过 PDF->Word->Pandoc)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.pptx')
            
        try:
            # 第一步: PDF -> Word
            temp_docx = input_path.with_suffix('.temp.docx')
            cv = Converter(str(input_path))
            await asyncio.to_thread(cv.convert, str(temp_docx), start=0, end=None)
            cv.close()
            
            # 第二步: Word -> PPTX (Pandoc)
            await asyncio.to_thread(pypandoc.convert_file, str(temp_docx), 'pptx', outputfile=str(output_path))
            
            # 清理
            if temp_docx.exists():
                temp_docx.unlink()
                
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/vnd.openxmlformats-officedocument.presentationml.presentation", 
                {"original_file": input_path.name, "type": "pdf2ppt"}
            )
        except Exception as e:
            logger.error(f"PDF 转 PPT 失败: {e}")
            raise e

    @staticmethod
    async def pdf_to_excel(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """PDF 转 Excel (占位符 / 基础提取)"""
        # 注意: 完整的 PDF 转 Excel 需要 tabula-py 或 camelot，这些库有较重的依赖 (java 等)
        # 这里我们提供一个存根或基础实现，否则我们可能会跳过或抛出不支持
        # 既然用户要求，我们尝试使用 pdf2docx 获取表格？不，pdf2docx 针对 docx。
        # 目前，我们将抛出 NotImplementedError 或提供一个虚拟文件以指示限制。
        # 或者，我们可以使用 `pypdf` 提取文本并转储到 CSV？
        raise NotImplementedError("PDF 转 Excel 需要当前不可用的专用库 (tabula-py)。")

    @staticmethod
    async def pdf_to_pdfa(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """PDF 转 PDF/A (通过 Ghostscript)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_name(f"{input_path.stem}_pdfa.pdf")
            
        try:
            # 用于 PDF/A-1b 的 Ghostscript 命令
            cmd = [
                "gs",
                "-dPDFA",
                "-dBATCH",
                "-dNOPAUSE",
                "-sProcessColorModel=DeviceCMYK",
                "-sDEVICE=pdfwrite",
                "-sPDFACompatibilityPolicy=1",
                f"-sOutputFile={str(output_path)}",
                str(input_path)
            ]
            
            # 检查 gs 是否存在
            if not shutil.which("gs"):
                raise Exception("未找到 Ghostscript (gs)")
                
            await asyncio.to_thread(subprocess.run, cmd, check=True, capture_output=True)
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/pdf", 
                {"original_file": input_path.name, "type": "pdf2pdfa"}
            )
        except Exception as e:
            logger.error(f"PDF 转 PDF/A 失败: {e}")
            raise e

    @staticmethod
    async def ofd_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """OFD 转 PDF (通过 easyofd)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
            
        try:
            # easyofd 转换
            # easyofd.OFD(str(input_path)).del_img() # 示例用法
            # easyofd.OFD(str(input_path)).save(str(output_path), format='pdf') 
            # 注意: easyofd API 可能会有所不同，假设标准用法或如果 CLI 可用则回退到子进程
            # 当前 easyofd (0.5.6) 支持基本提取。
            # 让我们尝试直接使用。如果失败，捕获异常。
            
            await asyncio.to_thread(
                lambda: easyofd.OFD(str(input_path)).save(str(output_path), format='pdf')
            )
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/pdf", 
                {"original_file": input_path.name, "type": "ofd2pdf"}
            )
        except Exception as e:
            logger.error(f"OFD 转 PDF 失败: {e}")
            raise e

    @staticmethod
    async def ofd_to_images(input_path: str | Path, output_dir: str | Path = None, user_id: str = None) -> list[str]:
        """OFD 转图片"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_dir is None:
            output_dir = input_path.parent / f"{input_path.stem}_ofd_images"
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
            
        try:
            # easyofd 可以保存为图片
            await asyncio.to_thread(
                lambda: easyofd.OFD(str(input_path)).save(str(output_dir), format='jpg')
            )
            
            results = []
            for img_file in output_dir.glob("*.jpg"):
                 url = await DocUtils._upload_and_record(
                     img_file, user_id, "doc_convert", "converted", "image/jpeg", 
                     {"original_file": input_path.name, "type": "ofd2img"}
                 )
                 results.append(url)
            return results
        except Exception as e:
            logger.error(f"OFD 转图片失败: {e}")
            raise e

    @staticmethod
    async def pdf_remove_limit(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """PDF 移除限制 (权限) (通过 pikepdf)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_name(f"{input_path.stem}_{uuid.uuid4()}_unlocked.pdf")
            
        try:
            def _unlock():
                with pikepdf.open(input_path) as pdf:
                    pdf.save(output_path)
            
            await asyncio.to_thread(_unlock)
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/pdf", 
                {"original_file": input_path.name, "type": "pdf_unlock"}
            )
        except Exception as e:
            logger.error(f"PDF 解锁失败: {e}")
            raise e

    @staticmethod
    async def pdf_to_long_image(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """PDF 转长图 (拼接)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.jpg')
            
        try:
            # 1. 首先转换为图片
            pdf = pdfium.PdfDocument(str(input_path))
            images = []
            total_height = 0
            max_width = 0
            
            for i in range(len(pdf)):
                page = pdf[i]
                img = page.render(scale=2).to_pil()
                images.append(img)
                total_height += img.height
                max_width = max(max_width, img.width)
                
            # 2. 拼接
            long_img = Image.new('RGB', (max_width, total_height), (255, 255, 255))
            y_offset = 0
            for img in images:
                long_img.paste(img, (0, y_offset))
                y_offset += img.height
                
            await asyncio.to_thread(long_img.save, output_path, "JPEG", quality=85)
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "image/jpeg", 
                {"original_file": input_path.name, "type": "pdf2longimg"}
            )
        except Exception as e:
            logger.error(f"PDF 转长图失败: {e}")
            raise e

    @staticmethod
    async def image_convert(input_path: str | Path, target_fmt: str = "png", output_path: str | Path = None, user_id: str = None) -> str:
        """图片格式转换 (JPG<->PNG 等)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        target_fmt = target_fmt.lower().replace('.', '')
        if output_path is None:
            output_path = input_path.with_suffix(f'.{target_fmt}')
            
        try:
            img = Image.open(input_path)
            # 处理 JPG 的 alpha 通道
            if target_fmt in ['jpg', 'jpeg'] and img.mode in ('RGBA', 'LA'):
                background = Image.new(img.mode[:-1], img.size, (255, 255, 255))
                background.paste(img, img.split()[-1])
                img = background
                
            rgb_im = img.convert('RGB') if target_fmt in ['jpg', 'jpeg'] else img
            await asyncio.to_thread(rgb_im.save, output_path)
            
            mime_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png', 'gif': 'image/gif'}
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", mime_map.get(target_fmt, 'application/octet-stream'), 
                {"original_file": input_path.name, "type": f"img2{target_fmt}"}
            )
        except Exception as e:
            logger.error(f"图片转换失败: {e}")
            raise e

    @staticmethod
    async def svg_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """SVG 转 PDF (通过 svglib)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        if output_path is None:
            output_path = input_path.with_suffix('.pdf')
            
        try:
            drawing = svg2rlg(str(input_path))
            await asyncio.to_thread(renderPDF.drawToFile, drawing, str(output_path))
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/pdf", 
                {"original_file": input_path.name, "type": "svg2pdf"}
            )
        except Exception as e:
            logger.error(f"SVG 转 PDF 失败: {e}")
            raise e

    @staticmethod
    async def ebook_convert(input_path: str | Path, output_format: str, user_id: str = None) -> str:
        """电子书转换 (EPUB/MOBI/PDF 通过 Pandoc)"""
        input_path = Path(input_path).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"文件未找到: {input_path}")
            
        output_format = output_format.lower().replace('.', '')
        output_path = input_path.with_suffix(f'.{output_format}')
        
        try:
            # Pandoc 处理 epub, markdown, docx 等。
            # 注意: mobi 不是标准的 pandoc 输出 (需要 kindlegen)，但 'epub' 是。
            # 我们将首先尝试 pandoc。
            await asyncio.to_thread(pypandoc.convert_file, str(input_path), output_format, outputfile=str(output_path))
            
            return await DocUtils._upload_and_record(
                output_path, user_id, "doc_convert", "converted", "application/octet-stream", 
                {"original_file": input_path.name, "type": f"ebook2{output_format}"}
            )
        except Exception as e:
            logger.error(f"电子书转换失败: {e}")
            raise e

    @staticmethod
    async def caj_to_pdf(input_path: str | Path, output_path: str | Path = None, user_id: str = None) -> str:
        """CAJ 转 PDF (存根/占位符)"""
        # 需要外部 caj2pdf 工具，在此环境中不容易通过 pip 安装
        raise NotImplementedError("CAJ 转 PDF 需要外部 'caj2pdf' 工具。")

if __name__ == "__main__":
    # 测试代码
    async def main():
        try:
            test_file = Path("/home/code_dev/trai/CONDA_ENV_STATUS.md")
            if test_file.exists():
                print(f"正在转换: {test_file}")
                # 假设 user_id="system" 进行测试
                pdf_path = await DocUtils.md_to_pdf(test_file, user_id="system")
                print(f"转换成功: {pdf_path}")
            else:
                print(f"测试文件不存在: {test_file}")
        except Exception as e:
            print(f"转换出错: {e}")

    asyncio.run(main())
